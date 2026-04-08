"""
Export Service - Handle exporting data to multiple formats
Supports: CSV, Excel (XLSX), PDF
Integrated with Celery for large exports
"""
import os
import io
import csv
import json
import hashlib
import logging
from datetime import datetime
from decimal import Decimal
from uuid import uuid4
from typing import Dict, List, Any, Optional, Tuple

from django.conf import settings
from django.utils import timezone
from django.db.models import QuerySet
from django.template.loader import render_to_string

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for handling data export with features:
    - Export CSV/Excel/PDF
    - Supports multiple data types (transactions, accounts, budgets, etc.)
    - Automatic file aggregation and cleanup
    - Integrated async with Celery
    """
    
    # Export configuration
    EXPORT_DIR = os.path.join(settings.MEDIA_ROOT, 'exports')
    MAX_SYNC_ROWS = 1000  # Row limit for sync export, exceeding triggers async
    FILE_EXPIRY_HOURS = 24  # Export file expires after 24h
    
    # Mapping of exportable data types
    EXPORTABLE_TYPES = {
        'transactions': {
            'title': 'Transactions',
            'fields': ['transaction_id', 'account_name', 'category_name', 'amount', 
                      'transaction_type', 'transaction_date', 'description', 'note', 'location'],
            'headers': ['Trans ID', 'Account', 'Category', 'Amount', 
                       'Type', 'Trans Date', 'Description', 'Note', 'Location'],
        },
        'accounts': {
            'title': 'Accounts',
            'fields': ['account_id', 'account_name', 'account_type', 'balance', 
                      'currency', 'bank_name', 'account_number'],
            'headers': ['Acct ID', 'Acct Name', 'Acct Type', 'Balance', 
                       'Currency', 'Bank', 'Account Number'],
        },
        'budgets': {
            'title': 'Budgets',
            'fields': ['budget_id', 'budget_name', 'category_name', 'amount', 
                      'period', 'start_date', 'end_date', 'spent_amount'],
            'headers': ['Budget ID', 'Budget Name', 'Category', 'Limit', 
                       'Period', 'Start Date', 'End Date', 'Spent'],
        },
        'debts': {
            'title': 'Debts',
            'fields': ['debt_id', 'debt_type', 'person_name', 'amount', 
                      'remaining_amount', 'interest_rate', 'start_date', 'due_date', 'status'],
            'headers': ['Debt ID', 'Debt Type', 'Related Person', 'Amount', 
                       'Remaining', 'Interest Rate (%)', 'Start Date', 'Due Date', 'Status'],
        },
        'savings': {
            'title': 'Savings Goals',
            'fields': ['goal_id', 'goal_name', 'target_amount', 'current_amount', 
                      'target_date', 'priority', 'status'],
            'headers': ['ID', 'Goal Name', 'Target', 'Current', 
                       'Target Date', 'Priority', 'Status'],
        },
    }

    # ==================== CORE EXPORT METHODS ====================
    
    @classmethod
    def prepare_export_data(cls, data_type: str, queryset: QuerySet) -> Tuple[List[Dict], List[str], List[str]]:
        """
        Prepare data for export from queryset.
        Returns: (data_list, fields, headers)
        """
        if data_type not in cls.EXPORTABLE_TYPES:
            raise ValueError(f"Data type '{data_type}' is not supported for export")
        
        config = cls.EXPORTABLE_TYPES[data_type]
        fields = config['fields']
        headers = config['headers']
        
        data_list = []
        for obj in queryset:
            row = {}
            for field in fields:
                value = None
                
                # Handle special related fields
                if field == 'account_name':
                    value = obj.account.account_name if hasattr(obj, 'account') and obj.account else ''
                elif field == 'category_name':
                    value = obj.category.category_name if hasattr(obj, 'category') and obj.category else ''
                elif field == 'spent_amount' and hasattr(obj, 'spent_amount'):
                    # For budgets with annotation
                    value = float(obj.spent_amount) if obj.spent_amount else 0
                else:
                    # Get directly from object
                    value = getattr(obj, field, '')
                
                # Convert data types
                if isinstance(value, Decimal):
                    value = float(value)
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(value, 'strftime'):  # Date object
                    value = value.strftime('%Y-%m-%d')
                
                row[field] = value
            data_list.append(row)
        
        return data_list, fields, headers
    
    @classmethod
    def export_to_csv(cls, data_type: str, queryset: QuerySet, user) -> str:
        """
        Export data to CSV.
        Returns: Path to created CSV file
        """
        data_list, fields, headers = cls.prepare_export_data(data_type, queryset)
        
        # Create export directory if not exists
        user_export_dir = os.path.join(cls.EXPORT_DIR, str(user.user_id))
        os.makedirs(user_export_dir, exist_ok=True)
        
        # Create unique filename
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{data_type}_{timestamp}.csv"
        filepath = os.path.join(user_export_dir, filename)
        
        # Write CSV with UTF-8 BOM so Excel reads Vietnamese correctly
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            # Write formatted headers
            header_dict = dict(zip(fields, headers))
            writer.writerow(header_dict)
            writer.writerows(data_list)
        
        logger.info(f"[EXPORT] CSV exported: {filepath} ({len(data_list)} rows)")
        return filepath

    @classmethod
    def export_to_excel(cls, data_type: str, queryset: QuerySet, user) -> str:
        """
        Export data to Excel with formatting.
        Returns: Path to created Excel file
        """
        data_list, fields, headers = cls.prepare_export_data(data_type, queryset)
        config = cls.EXPORTABLE_TYPES[data_type]
        
        # Create directory
        user_export_dir = os.path.join(cls.EXPORT_DIR, str(user.user_id))
        os.makedirs(user_export_dir, exist_ok=True)
        
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{data_type}_{timestamp}.xlsx"
        filepath = os.path.join(user_export_dir, filename)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = config['title']
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write report title
        ws.merge_cells('A1:' + chr(64 + len(headers)) + '1')
        title_cell = ws['A1']
        title_cell.value = f"BÁO CÁO {config['title'].upper()}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:' + chr(64 + len(headers)) + '2')
        date_cell = ws['A2']
        date_cell.value = f"Exported on: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal='center')
        
        # Ghi headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Ghi data
        for row_idx, row_data in enumerate(data_list, 5):
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ''))
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                
                # Format amounts
                if 'amount' in field or 'balance' in field:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filepath)
        logger.info(f"[EXPORT] Excel exported: {filepath} ({len(data_list)} rows)")
        return filepath

    @classmethod
    def export_to_pdf(cls, data_type: str, queryset: QuerySet, user) -> str:
        """
        Export data to PDF using WeasyPrint.
        Returns: Path to created PDF file
        """
        try:
            from weasyprint import HTML, CSS
        except ImportError:
            raise ImportError("WeasyPrint is not installed. Run: pip install weasyprint")
        
        data_list, fields, headers = cls.prepare_export_data(data_type, queryset)
        config = cls.EXPORTABLE_TYPES[data_type]
        
        # Create directory
        user_export_dir = os.path.join(cls.EXPORT_DIR, str(user.user_id))
        os.makedirs(user_export_dir, exist_ok=True)
        
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{data_type}_{timestamp}.pdf"
        filepath = os.path.join(user_export_dir, filename)
        
        # Create HTML content
        html_content = cls._generate_pdf_html(config['title'], headers, fields, data_list, user)
        
        # CSS cho PDF
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 1cm;
            }
            body {
                font-family: "DejaVu Sans", Arial, sans-serif;
                font-size: 10px;
            }
            h1 {
                color: #2c3e50;
                text-align: center;
                margin-bottom: 5px;
            }
            .meta {
                text-align: center;
                color: #7f8c8d;
                margin-bottom: 20px;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }
            th {
                background-color: #3498db;
                color: white;
                padding: 8px;
                text-align: left;
                font-weight: bold;
            }
            td {
                padding: 6px 8px;
                border-bottom: 1px solid #ddd;
            }
            tr:nth-child(even) {
                background-color: #f9f9f9;
            }
            .amount {
                text-align: right;
                font-family: monospace;
            }
            .footer {
                margin-top: 20px;
                text-align: center;
                color: #95a5a6;
                font-size: 9px;
            }
        ''')
        
        # Create PDF
        HTML(string=html_content).write_pdf(filepath, stylesheets=[css])
        logger.info(f"[EXPORT] PDF exported: {filepath} ({len(data_list)} rows)")
        return filepath

    @classmethod
    def _generate_pdf_html(cls, title: str, headers: List[str], fields: List[str], 
                           data_list: List[Dict], user) -> str:
        """Generate HTML template for PDF"""
        rows_html = ""
        for row in data_list:
            row_html = "<tr>"
            for field in fields:
                value = row.get(field, '')
                css_class = 'amount' if 'amount' in field or 'balance' in field else ''
                if isinstance(value, (int, float)):
                    value = f"{value:,.2f}"
                row_html += f"<td class='{css_class}'>{value}</td>"
            row_html += "</tr>"
            rows_html += row_html
        
        headers_html = "".join([f"<th>{h}</th>" for h in headers])
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Report {title}</title>
        </head>
        <body>
            <h1>BÁO CÁO {title.upper()}</h1>
            <div class="meta">
                User: {user.full_name} | Exported on: {timezone.now().strftime('%d/%m/%Y %H:%M')}
            </div>
            <table>
                <thead>
                    <tr>{headers_html}</tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
            <div class="footer">
                Auto-generated by Expense Management System | Total: {len(data_list)} records
            </div>
        </body>
        </html>
        """
        return html

    # ==================== UTILITY METHODS ====================
    
    @classmethod
    def get_export_file_url(cls, filepath: str) -> str:
        """Convert file path to download URL"""
        relative_path = os.path.relpath(filepath, settings.MEDIA_ROOT)
        return f"{settings.MEDIA_URL}{relative_path}"
    
    @classmethod
    def cleanup_old_exports(cls, hours: int = None):
        """
        Delete export files older than the specified time.
        Called by Celery beat periodically.
        """
        hours = hours or cls.FILE_EXPIRY_HOURS
        cutoff_time = timezone.now() - timezone.timedelta(hours=hours)
        deleted_count = 0
        
        if not os.path.exists(cls.EXPORT_DIR):
            return 0
        
        for root, dirs, files in os.walk(cls.EXPORT_DIR):
            for filename in files:
                filepath = os.path.join(root, filename)
                file_mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
                file_mtime = timezone.make_aware(file_mtime)
                
                if file_mtime < cutoff_time:
                    try:
                        os.remove(filepath)
                        deleted_count += 1
                        logger.debug(f"[CLEANUP] Deleted old export: {filepath}")
                    except Exception as e:
                        logger.error(f"[CLEANUP] Failed to delete {filepath}: {e}")
        
        logger.info(f"[CLEANUP] Deleted {deleted_count} old export files")
        return deleted_count

    @classmethod
    def should_use_async(cls, queryset: QuerySet) -> bool:
        """Check whether async export should be used"""
        return queryset.count() > cls.MAX_SYNC_ROWS

    @classmethod
    def list_user_exports(cls, user, limit: int = 20) -> List[Dict]:
        """List recent export files for a user"""
        user_export_dir = os.path.join(cls.EXPORT_DIR, str(user.user_id))
        
        if not os.path.exists(user_export_dir):
            return []
        
        exports = []
        for filename in os.listdir(user_export_dir):
            filepath = os.path.join(user_export_dir, filename)
            stat = os.stat(filepath)
            exports.append({
                'filename': filename,
                'url': cls.get_export_file_url(filepath),
                'size': stat.st_size,
                'size_human': cls._format_file_size(stat.st_size),
                'created_at': datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
        
        # Sort by newest first
        exports.sort(key=lambda x: x['created_at'], reverse=True)
        return exports[:limit]

    @staticmethod
    def _format_file_size(size: int) -> str:
        """Format file size to human-readable format"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.1f} {unit}"
            size /= 1024
        return f"{size:.1f} TB"
